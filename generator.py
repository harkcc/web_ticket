from numpy import add
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
import os
import traceback
import json
from datetime import datetime
from db_connector import MongoDBConnector
from io import BytesIO
from openpyxl.utils import get_column_letter
import re


# 记得切环境


class ProcessingError(Exception):
    """处理错误的自定义异常类"""
    pass


def template_handler(keyword):
    """模板处理器装饰器"""
    def decorator(func):
        func._template_keyword = keyword
        return func
    return decorator


class InvoiceGenerator:
    def __init__(self, upload_folder, output_folder, db_connector=None, image_folder=None):
        """
        初始化发票生成器
        :param upload_folder: 上传文件夹路径
        :param output_folder: 输出文件夹路径
        :param db_connector: 数据库连接器（可选）
        :param image_folder: 图片文件夹路径（可选）
        """
        self.db_connector = db_connector if db_connector else MongoDBConnector()
        self.upload_folder = upload_folder
        self.output_folder = output_folder
        # 获取当前文件所在的目录
        current_dir = os.path.dirname(os.path.abspath(__file__))
        self.image_folder = os.path.join(current_dir, '产品图片(1)')  # 图片文件夹路径
        
        # 定义模板配置，只列出不需要编码的模板
        self.template_config = {
            "依诺达": {"requires_code": False}  # 不需要编码的模板
        }
        
        # 初始化模板处理器字典
        self._template_handlers = {}
        # 注册所有带有_template_keyword属性的方法
        print("开始注册模板处理器...")
        for name in dir(self):
            method = getattr(self, name)
            if hasattr(method, '_template_keyword'):
                keyword = method._template_keyword
                self._template_handlers[keyword] = method
                print(f"注册模板处理器: {name} -> {keyword}")
        print(f"已注册的模板处理器: {list(self._template_handlers.keys())}")

    @template_handler("叮铛卡航限时达")
    def _fill_dingdang_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """
        填充叮铛卡航限时达模板
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                sheet = wb['模板']  # 获取模板工作表

                print("开始写入模版信息")

                # 定义样式信息
                style_info = {
                    'font': Font(name='Arial', size=10),
                    'border': Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin')),
                    'alignment': Alignment(horizontal='center', vertical='center')
                }

                # 在第一行B列填充编码
                if code:
                    cell = sheet.cell(row=1, column=2)  # B列是第2列
                    cell.value = code
                    cell.font = Font(name='Arial', size=9)

                # 如果有地址信息，填充到相应的单元格
                if address_info:
                    address_info_detail = address_info['address_info']
                    try:
                        # 填充收件人信息
                       
                        # 填充地址信息
                        address_parts = []
                        if 'name' in address_info_detail:
                            cell = sheet.cell(row=4, column=2)  # B2单元格
                            cell.value = address_info_detail['name']
                            cell = sheet.cell(row=3, column=2)  # B2单元格
                            cell.value = address_info_detail['name']
                            address_parts.append(address_info_detail['name'])
                        if 'addressLine1' in address_info_detail:
                            address_parts.append(address_info_detail['addressLine1'])
                        if 'city' in address_info_detail:
                            address_parts.append(address_info_detail['city'])
                        if 'stateOrProvinceCode' in address_info_detail:
                            address_parts.append(address_info_detail['stateOrProvinceCode'])
                        if 'postalCode' in address_info_detail:
                            address_parts.append(address_info_detail['postalCode'])
                        if 'countryCode' in address_info_detail:
                            address_parts.append(address_info_detail['countryCode'])

                        # 填充地址信息
                        if 'addressLine1' in address_info_detail:
                            cell = sheet.cell(row=6, column=2)  # B3单元格
                            cell.value = address_info_detail['addressLine1']

                        # 城市
                        if 'city' in address_info_detail:
                            cell = sheet.cell(row=8, column=2)  # B4单元格
                            cell.value = address_info_detail['city']

                        #邮政编码
                        if 'postalCode' in address_info_detail:
                            cell = sheet.cell(row=10, column=2)  # B6单元格
                            cell.value = address_info_detail['postalCode']

                        #国家代码
                        if 'countryCode' in address_info_detail:
                            cell = sheet.cell(row=11, column=2)  # B7单元格
                            cell.value = address_info_detail['countryCode']

                        if address_parts:
                            cell = sheet.cell(row=6, column=2)  # B3单元格
                            cell.value = ', '.join(address_parts)
                    except Exception as e:
                        print(f"填充地址信息时发生错误: {str(e)}")

                try:
                    total_boxes = len(box_data.keys())
                    cell = sheet.cell(row=16, column=2)  # 在第7行B列填充箱数
                    cell.value = str(total_boxes)
                    cell.font = Font(name='Arial', size=9)
                except Exception as e:
                    print(f"填充箱数时发生错误: {str(e)}")

                # 检查所有产品的电磁属性
                has_electric = False
                has_magnetic = False
                for box in box_data.values():
                    for item in box.items:
                        product_info = self._get_product_info(item.msku, db)
                        if product_info:
                            if product_info.get('electrified', '') == '是':
                                has_electric = True
                            if product_info.get('magnetic', '') == '是':
                                has_magnetic = True
                            if has_electric and has_magnetic:
                                break
                    if has_electric and has_magnetic:
                        break

                # 在表格顶部添加电磁属性标记
                if has_electric:
                    cell = sheet.cell(row=1, column=6)  # F列第1行
                    cell.value = "是"
                    cell.font = Font(name='Arial', size=9)
                
                if has_magnetic:
                    cell = sheet.cell(row=2, column=6)  # F列第2行
                    cell.value = "是"
                    cell.font = Font(name='Arial', size=9)

                # 填充数据
                row_num = 18  # 从第18行开始填充
                index = 1    # 添加序号计数器，从1开始

                # 遍历每个箱子

                sorted_boxes = sorted(box_data.items(), key=lambda x: int(x[0]))
                
                # 遍历排序后的箱子
                for box_number, box in sorted_boxes:
                    print(f"处理箱子 {box_number}")

                    # 遍历箱子中的每个产品
                    for item in box.items:
                        # 从数据库获取产品信息
                        product_info = self._get_product_info(item.msku, db)
                        # 处理产品信息为None的情况
                        if product_info is None:
                            print(f"警告: 未找到产品 {item.msku} 的信息")
                            price = 0
                            total_price = 0
                        else:
                            price = product_info.get('price', 0)
                            total_price = float(price) * item.box_quantities.get(box_number, 0) if price else 0
                            item.product_name = product_info.get('cn_name', item.product_name)
                        
                        # 设置单元格值和样式ç
                        cell_data = [
                            (1, box_number),                    # 货箱编号 (A列)
                            (2, box.weight if box.weight is not None else ""),  # 重量 (B列)
                            (3,product_info.get('en_name', '') if product_info else ''),  # 链接 (D列)
                            (4, product_info.get('cn_name', '') if product_info else ''),  # 链接 (D列)
                            (5, product_info.get('price', '') if product_info else ''),   # 仅在总价格大于0时填入
                            (6, item.box_quantities.get(box_number, 0)),  # 数量 (F列)
                            (7, str(product_info.get('material_en', '')+'/'+product_info.get('material_cn', '')) if product_info else ''),  # 材料 (D列) 
                            (8, product_info.get('hs_code', '') if product_info else ''),  # HS编码 (G列)
                            (9, str(product_info.get('usage_en', '')+'/'+product_info.get('usage_cn', '' ))if product_info else ''),    # 用途 (H列)
                            (10, product_info.get('brand', '') if product_info else ''),    # 品牌 (I列)
                            (11, product_info.get('model', '') if product_info else ''),   # 型号 (J列)
                            (12, product_info.get('link', '') if product_info else ''),
                            (14, ''),  # 图片列 (N列)
                            (15, total_price if total_price > 0 else ""),  # 仅在总价格大于0时填入
                            (17, box.length if box.length is not None else ""),  # 长度 (Q列)
                            (18, box.width if box.width is not None else ""),    # 宽度 (R列)
                            (19, box.height if box.height is not None else "")   # 高度 (S列)
                        ]

                        # 批量设置单元格值和样式
                        for column, value in cell_data:
                            self._set_cell_value(sheet, row_num, column, value, style_info)

                        # 插入产品图片
                        if item.msku and hasattr(self, 'image_folder'):
                            try:
                                image_cell = f"N{row_num}"  # 图片列（第14列）
                                # self.insert_product_image(sheet, image_cell, item.msku, self.image_folder)
                                self.insert_original_product_image(sheet, image_cell, item.msku, self.image_folder)
                            except Exception as e:
                                print(f"插入图片时发生错误: {str(e)}")

                        row_num += 1

            except Exception as e:
                print(f"填充模板时发生错误: {str(e)}")
                raise

    @template_handler("顺丰")
    def _fill_sf_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """填充顺丰模板"""
        """
        填充顺丰模板
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                sheet = wb['Sheet1']  # 获取模板工作表
                # box_Reference_id = ''  # 在方法开始时就初始化
                print("开始写入模版信息")

                # 定义样式信息
                style_info = {
                    'font': Font(name='Arial', size=10),
                    'border': Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin')),
                    'alignment': Alignment(horizontal='center', vertical='center')
                }

                self.unmerge_cells_in_range(sheet, 2, 4, 2, 9)

                # 如果有地址信息，填充到相应的单元格
                if address_info:
                    address_info_detail = address_info['address_info']
                    # if 'seller_info' in address_info:
                    #     box_Reference_id = address_info['seller_info']['amazonReferenceId']
                    try:
                        # 填充收件人信息
                       
                        # 填充地址信息
                        address_parts = []
                        if 'name' in address_info_detail:
                            cell = sheet.cell(row=2, column=2)  
                            cell.value = address_info_detail['name']

                            cell = sheet.cell(row=3, column=2)  
                            cell.value = address_info_detail['name']
                            address_parts.append(address_info_detail['name'])
                        if 'addressLine1' in address_info_detail:
                            address_parts.append(address_info_detail['addressLine1'])
                        if 'city' in address_info_detail:
                            address_parts.append(address_info_detail['city'])
                        if 'stateOrProvinceCode' in address_info_detail:
                            address_parts.append(address_info_detail['stateOrProvinceCode'])
                        if 'postalCode' in address_info_detail:
                            address_parts.append(address_info_detail['postalCode'])
                        if 'countryCode' in address_info_detail:
                            address_parts.append(address_info_detail['countryCode'])

                        if address_parts:
                            cell = sheet.cell(row=4, column=2)  
                            cell.value = ', '.join(address_parts)
  
                    except Exception as e:
                        print(f"填充地址信息时发生错误: {str(e)}")

                # 填充数据
                row_num = 12  
                index = 1    # 添加序号计数器，从1开始

                # 将box_data按箱号排序
                sorted_boxes = sorted(box_data.items(), key=lambda x: int(x[0]))
                
                # 遍历排序后的箱子
                for box_number, box in sorted_boxes:
                    print(f"处理箱子 {box_number}")

                    # 遍历箱子中的每个产品
                    for item in box.items:
                        # 从数据库获取产品信息
                        product_info = self._get_product_info(item.msku, db)
                        # print(f"产品信息：{product_info}")
                        # price = product_info.get('price', 0)
                        # total_price = float(price) * item.box_quantities.get(box_number, 0) if price else 0
                        price = 0
                        total_price = 0
                        # if product_info:
                        #     item.product_name = product_info.get('cn_name', item.product_name)


                        # print(f"产品信息：{product_info}")
                        if product_info is not None:
                            item.product_name = product_info.get('cn_name', item.product_name)
                            print(f"产品信息：{product_info}")
                        else:
                        # 处理未找到产品信息的情况
                            print(f"未找到产品信息，MSKU: {item.msku}")
                            item.product_name = "需要补数据"  # 可以设置一个默认值
                        
                        # box_number_str = code+f"{box_number:05d}" 
                        box_number_str = code+'U00000'+str(box_number)
                        # Reference_id = ''  # 初始化为None
                        # if box_Reference_id:
                        #     Reference_id = box_Reference_id
                        
                        # 设置单元格值和样式
                        cell_data = [
                            # 基本信息
                            (1, box_number_str),                                   # 货箱编号
                            # (2, Reference_id),                                    # 参考编号
                            (3, item.msku),                                       # 商品编码
                            # 产品名称信息
                            (4, product_info.get('en_name', '') if product_info else ''),                 # 英文名称
                            (5, product_info.get('cn_name', '') if product_info else ''),                 # 中文名称
                            (6, product_info.get('brand', '') if product_info else ''),                   # 品牌
                            (7, product_info.get('model', '') if product_info else ''),                   # 型号
                            
                            # 产品材料和用途
                            (8, product_info.get('material_cn', '') if product_info else ''),             # 中文材料
                            (9, product_info.get('material_en', '') if product_info else ''),             # 英文材料
                            (10, str(product_info.get('usage_en', '') + 
                                   product_info.get('usage_cn', '')) if product_info else ''),            # 用途
                            (11, '纸箱'),                                         # 包装类型
                            
                            # 产品规格信息
                            (12, product_info.get('hs_code', '') if product_info else ''),                # HS编码
                            (13, item.box_quantities.get(box_number, 0)),         # 数量
                            
                            # 箱体信息
                            (16, box.length if box.length is not None else ""),   # 长度
                            (17, box.width if box.width is not None else ""),     # 宽度
                            (18, box.height if box.height is not None else ""),   # 高度
                            (19, box.weight if box.weight is not None else ""),   # 重量
                            
                            # 其他信息
                            (20, product_info.get('link', '') if product_info else ''),                   # 链接
                            (21, '')                                              # 图片占位
                        ]
                        # 批量设置单元格值和样式
                        for column, value in cell_data:
                            self._set_cell_value(sheet, row_num, column, value, style_info)

                        sheet.row_dimensions[row_num].height = sheet.row_dimensions[12].height
                        
                        # 插入产品图片
                        if item.msku and hasattr(self, 'image_folder'):
                            try:
                                image_cell = f"U{row_num}"  # 图片列（第14列）
                                # self.insert_product_image(sheet, image_cell, item.msku, self.image_folder)
                                self.insert_original_product_image(sheet, image_cell, item.msku, self.image_folder)
                            except Exception as e:
                                print(f"插入图片时发生错误: {str(e)}")

                        row_num += 1
                
                self.merge_cells_in_range(sheet, 2, 2, 2, 9)
                self.merge_cells_in_range(sheet, 3, 3, 2, 9)
                self.merge_cells_in_range(sheet, 4, 4, 2, 9)

            except Exception as e:
                print(f"填充模板时发生错误: {str(e)}")
                raise

    @template_handler("依诺达")
    def _fill_ynd_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """
        填充叮铛卡航限时达模板
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                sheet = wb['模板']  # 获取模板工作表
                self.unmerge_cells_in_range(sheet, 15, 15, 2, 4)
                self.unmerge_cells_in_range(sheet, 1, 1, 6, 8)
                self.unmerge_cells_in_range(sheet, 2, 2, 6, 8)
                print("开始写入模版信息")

                # 定义样式信息
                style_info = {
                    'font': Font(name='Arial', size=10),
                    'border': Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin')),
                    'alignment': Alignment(horizontal='center', vertical='center')
                }

                try:
                    total_boxes = len(box_data.keys())
                    cell = sheet.cell(row=15, column=2)  #填充箱数
                    cell.value = str(total_boxes)
                    cell.font = Font(name='Arial', size=9)
                except Exception as e:
                    print(f"填充箱数时发生错误: {str(e)}")

                # 检查所有产品的电磁属性
                has_electric = False
                has_magnetic = False
                for box in box_data.values():
                    for item in box.items:
                        product_info = self._get_product_info(item.msku, db)
                        if product_info:
                            if product_info.get('electrified', '') == '是':
                                has_electric = True
                            if product_info.get('magnetic', '') == '是':
                                has_magnetic = True
                            if has_electric and has_magnetic:
                                break
                    if has_electric and has_magnetic:
                        break

                # 在表格顶部添加电磁属性标记
                if has_electric:
                    cell = sheet.cell(row=1, column=6)  # F列第1行
                    cell.value = "是"
                    cell.font = Font(name='Arial', size=9)
                
                if has_magnetic:
                    cell = sheet.cell(row=2, column=6)  # F列第2行
                    cell.value = "是"
                    cell.font = Font(name='Arial', size=9)

                # 填充数据
                row_num = 17  # 从第18行开始填充
                index = 1    # 添加序号计数器，从1开始
                row_height = sheet.row_dimensions[17].height
                sheet.column_dimensions['Q'].width = row_height/4
                
                # 遍历每个箱子

                sorted_boxes = sorted(box_data.items(), key=lambda x: int(x[0]))
                
                # 遍历排序后的箱子
                for box_number, box in sorted_boxes:
                    print(f"处理箱子 {box_number}")

                    # 遍历箱子中的每个产品
                    for item in box.items:
                        # 从数据库获取产品信息
                        product_info = self._get_product_info(item.msku, db)
                        # 处理产品信息为None的情况
                        if product_info is None:
                            print(f"警告: 未找到产品 {item.msku} 的信息")
                            price = 0
                            total_price = 0
                        else:
                            price = product_info.get('price', 0)
                            total_price = float(price) * item.box_quantities.get(box_number, 0) if price else 0
                            item.product_name = product_info.get('cn_name', item.product_name)
                        
                        # 设置单元格值和样式ç
                        cell_data = [
                            (1, box_number),                    # 货箱编号 (A列)
                            (2, box.weight if box.weight is not None else ""),  # 重量 
                            (3, box.length if box.length is not None else ""),  # 长度 
                            (4, box.width if box.width is not None else ""),    # 宽度 
                            (5, box.height if box.height is not None else "") ,  # 高度 
                            (6, item.msku),                    
                            (7,product_info.get('en_name', '') if product_info else ''),  # 链接 (D列)
                            (8, product_info.get('cn_name', '') if product_info else ''),  # 链接 (D列)
                            # (5, product_info.get('price', '') if product_info else ''),   # 仅在总价格大于0时填入
                            (10, item.box_quantities.get(box_number, 0)),  # 数量 (F列)
                            (11, str(product_info.get('material_en', '')+'/'+product_info.get('material_cn', '')) if product_info else ''),  # 材料 (D列) 
                            (13, product_info.get('hs_code', '') if product_info else ''),  # HS编码 (G列)
                            (12, str(product_info.get('usage_en', '')+'/'+product_info.get('usage_cn', '' ))if product_info else ''),    # 用途 (H列)
                            (14, product_info.get('brand', '') if product_info else ''),    # 品牌 (I列)
                            (15, product_info.get('model', '') if product_info else ''),   # 型号 (J列)
                            (16, product_info.get('link', '') if product_info else ''),
                            (17, ''),  # 图片列 (N列)
                            # (15, total_price if total_price > 0 else ""),  # 仅在总价格大于0时填入
                          
                        ]

                        # 批量设置单元格值和样式
                        for column, value in cell_data:
                            self._set_cell_value(sheet, row_num, column, value, style_info)
                        sheet.row_dimensions[row_num].height = row_height
                       
                        # 插入产品图片
                        if item.msku and hasattr(self, 'image_folder'):
                            try:
                                image_cell = f"Q{row_num}"
                                # self.insert_product_image(sheet, image_cell, item.msku, self.image_folder)
                                self.insert_original_product_image(sheet, image_cell, item.msku, self.image_folder)
                            except Exception as e:
                                print(f"插入图片时发生错误: {str(e)}")

                        row_num += 1

                self.merge_cells_in_range(sheet, 15, 15, 2, 4)
                self.merge_cells_in_range(sheet, 1, 1, 6, 8)
                self.merge_cells_in_range(sheet, 2, 2, 6, 8)

            except Exception as e:
                print(f"填充模板时发生错误: {str(e)}")
                raise

    @template_handler("叮铛(美洲)")
    def _fill_ddmz_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """
        填充叮铛卡航限时达模板
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                sheet = wb['清关发票']  # 获取模板工作表

                print("开始写入模版信息")

                # 定义样式信息
                style_info = {
                    'font': Font(name='Arial', size=10),
                    'border': Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin')),
                    'alignment': Alignment(horizontal='center', vertical='center')
                }

                self.unmerge_cells_in_range(sheet, 3, 3, 8, 15)
                self.unmerge_cells_in_range(sheet, 4, 4, 8, 15)
                self.unmerge_cells_in_range(sheet, 5, 5, 8, 11)
                self.unmerge_cells_in_range(sheet, 5, 5, 13, 15)
                self.unmerge_cells_in_range(sheet, 6, 6, 8, 15)
                self.unmerge_cells_in_range(sheet, 7, 7, 8, 11)
                self.unmerge_cells_in_range(sheet, 7, 7,13, 15)
                self.unmerge_cells_in_range(sheet, 8, 8,13, 15)

                # 在第一行B列填充编码
                if code:
                    cell = sheet.cell(row=3, column=2)  # B列是第2列
                    cell.value = code
                    cell.font = Font(name='Arial', size=9)

                # 如果有地址信息，填充到相应的单元格
                if address_info:
                    address_info_detail = address_info['address_info']
                    try:
                        # 填充收件人信息，这里收件人和
                        
                        # 填充地址信息
                        address_parts = []
                        if 'name' in address_info_detail:
                            cell = sheet.cell(row=3, column=8)  # B2单元格
                            cell.value = address_info_detail['name']

                            cell = sheet.cell(row=4, column=8)  # B2单元格
                            cell.value = address_info_detail['name']

                            cell = sheet.cell(row=8, column=13)  # B2单元格
                            cell.value = address_info_detail['name']
                            address_parts.append(address_info_detail['name'])
                        if 'addressLine1' in address_info_detail:
                            address_parts.append(address_info_detail['addressLine1'])
                        if 'city' in address_info_detail:
                            address_parts.append(address_info_detail['city'])
                        if 'stateOrProvinceCode' in address_info_detail:
                            address_parts.append(address_info_detail['stateOrProvinceCode'])
                        if 'postalCode' in address_info_detail:
                            address_parts.append(address_info_detail['postalCode'])
                        if 'countryCode' in address_info_detail:
                            address_parts.append(address_info_detail['countryCode'])

                        # 填充地址信息
                        # if 'addressLine1' in address_info_detail:
                        #     cell = sheet.cell(row=6, column=2)  # B3单元格
                        #     cell.value = address_info_detail['addressLine1']

                        # 城市
                        if 'city' in address_info_detail:
                            cell = sheet.cell(row=7, column=8)  # B4单元格
                            cell.value = address_info_detail['city']

                        #省份
                        if 'stateOrProvinceCode' in address_info_detail:
                            cell = sheet.cell(row=7, column=13)  # B5单元格
                            cell.value = address_info_detail['stateOrProvinceCode']

                        #邮政编码
                        if 'postalCode' in address_info_detail:
                            cell = sheet.cell(row=5, column=13)  # B6单元格
                            cell.value = address_info_detail['postalCode']

                        #国家代码
                        if 'countryCode' in address_info_detail:
                            cell = sheet.cell(row=5, column=8)  # B7单元格
                            cell.value = address_info_detail['countryCode']

                        if address_parts:
                            cell = sheet.cell(row=6, column=8)  # B3单元格
                            cell.value = ', '.join(address_parts)
                    except Exception as e:
                        print(f"填充地址信息时发生错误: {str(e)}")

                try:
                    total_boxes = len(box_data.keys())
                    cell = sheet.cell(row=3, column=6)  # 在第7行B列填充箱数
                    cell.value = str(total_boxes)
                    cell.font = Font(name='Arial', size=9)
                except Exception as e:
                    print(f"填充箱数时发生错误: {str(e)}")


                # 填充数据
                row_num = 12  # 从第18行开始填充
                index = 1    # 添加序号计数器，从1开始
                row_height = sheet.row_dimensions[12].height
                Reference_id = ''

                # 遍历每个箱子
                sorted_boxes = sorted(box_data.items(), key=lambda x: int(x[0]))
                
                # 遍历排序后的箱子
                for box_number, box in sorted_boxes:
                    print(f"处理箱子 {box_number}")

                    # 遍历箱子中的每个产品
                    for item in box.items:
                        # 从数据库获取产品信息
                        product_info = self._get_product_info(item.msku, db)
                        # 处理产品信息为None的情况
                        if product_info is None:
                            print(f"警告: 未找到产品 {item.msku} 的信息")
                            price = 0
                            total_price = 0
                        else:
                            price = product_info.get('price', 0)
                            total_price = float(price) * item.box_quantities.get(box_number, 0) if price else 0
                            item.product_name = product_info.get('cn_name', item.product_name)
                        Reference_id = address_info['address_info'].get('amazonReferenceId','')

                        # 设置单元格值和样式
                        cell_data = [
                            (1, code+f"{box_number:05d}"),                    
                            # (2, box.weight if box.weight is not None else ""),  
                            (2,Reference_id),  
                            (3,f"{box.length}*{box.width}*{box.height}"),  
                            (4,box_number), 
                            (5,box.weight), 
                            (6, product_info.get('hs_code', '') if product_info else ''),  
                            (8,product_info.get('en_name', '') if product_info else ''),  
                            (7, product_info.get('cn_name', '') if product_info else ''), 
                            (9, item.box_quantities.get(box_number, 0)),
                            (10,''),
                            (11, product_info.get('brand', '') if product_info else ''),    
                            (12, product_info.get('model', '') if product_info else ''),  
                            (13, str(product_info.get('material_cn', '')) if product_info else ''),  
                            (14, str(product_info.get('usage_cn', '')+product_info.get('usage_en', '' ))if product_info else ''),    # 用途 (H列)
                            (15, ''),  
                        ]

                        # 批量设置单元格值和样式
                        for column, value in cell_data:
                            self._set_cell_value(sheet, row_num, column, value, style_info)

                        sheet.row_dimensions[row_num].height = row_height

                        # 插入产品图片
                        if item.msku and hasattr(self, 'image_folder'):
                            try:
                                image_cell = f"O{row_num}"  # 图片列（第14列）
                                # self.insert_product_image(sheet, image_cell, item.msku, self.image_folder)
                                self.insert_original_product_image(sheet, image_cell, item.msku, self.image_folder)
                            except Exception as e:
                                print(f"插入图片时发生错误: {str(e)}")

                        row_num += 1
                
                self.merge_cells_in_range(sheet, 3, 3, 8, 15)
                self.merge_cells_in_range(sheet, 4, 4, 8, 15)
                self.merge_cells_in_range(sheet, 5, 5, 8, 11)
                self.merge_cells_in_range(sheet, 5, 5, 13, 15)
                self.merge_cells_in_range(sheet, 6, 6, 8, 15)
                self.merge_cells_in_range(sheet, 7, 7, 8, 11)
                self.merge_cells_in_range(sheet, 7, 7,13, 15)
                self.merge_cells_in_range(sheet, 8, 8,13, 15)
                
            except Exception as e:
                print(f"填充模板时发生错误: {str(e)}")
                raise

    @template_handler("UPS(美洲)")
    def _fill_ups_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """
        填充UPS美洲模板
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                sheet = wb['发票']  # 获取发票工作表
                
                # 定义样式信息
                style_info = {
                    'font': Font(name='Arial', size=10),
                    'border': Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin')),
                    'alignment': Alignment(horizontal='center', vertical='center')
                }


                # 解除合并单元格
                print("正在解除合并单元格...")
                # self.unmerge_cells_in_range(sheet, 13, 16, 1, 15)
                self.unmerge_cells_in_range(sheet, 4, 4, 1, 3)
                self.unmerge_cells_in_range(sheet, 7, 11, 1, 3)
                self.unmerge_cells_in_range(sheet, 7, 11, 4, 15)
                
                row_num = 13  # 从第13行开始填充数据
                total_quantity = 0
                total_amount = 0
                total_weight = 0
                
                
                if code:
                    cell = sheet.cell(row=4, column=1)  # B列是第2列
                    cell.value = f"运单号码:{code}"
                    cell.font = Font(name='Arial', size=12)

                    cell_another = sheet.cell(row=7, column=1)  # B列是第2列
                    cell_another.value = f"FBA编号:{code}"
                    cell_another.font = Font(name='Arial', size=12)
                
                  # 如果有地址信息，填充到相应的单元格
                if address_info:
                    address_info_detail = address_info['address_info']
                    try:

                        # 填充地址信息
                        address_parts = []
                        if 'name' in address_info_detail:
                            address_parts.append(address_info_detail['name'])
                        if 'addressLine1' in address_info_detail:
                            address_parts.append(address_info_detail['addressLine1'])
                        if 'city' in address_info_detail:
                            address_parts.append(address_info_detail['city'])
                        if 'stateOrProvinceCode' in address_info_detail:
                            address_parts.append(address_info_detail['stateOrProvinceCode'])
                        if 'postalCode' in address_info_detail:
                            address_parts.append(address_info_detail['postalCode'])
                        if 'countryCode' in address_info_detail:
                            address_parts.append(address_info_detail['countryCode'])

                        if address_parts:
                            cell = sheet.cell(row=7, column=4)  # B3单元格
                            cell.value = ', '.join(address_parts)
                            cell.font = Font(name='Arial', size=12)
                    except Exception as e:
                        print(f"填充地址信息时发生错误: {str(e)}")
            
                # 遍历每个箱子
                sorted_boxes = sorted(box_data.items(), key=lambda x: int(x[0]))
                row_height = sheet.row_dimensions[13].height
                
                # 遍历排序后的箱子
                for box_number, box in sorted_boxes:
                    if not box.items:
                        continue
                    
                    start_row = row_num  # 记录当前箱子的起始行
                    box_items_count = len(box.items)
                    
                    for item in box.items:
                        # 获取产品信息
                        product_info = self._get_product_info(item.msku, db)
                        # 处理产品信息为None的情况
                        if product_info is None:
                            print(f"警告: 未找到产品 {item.msku} 的信息")
                            price = 0
                            total_price = 0
                        else:
                            price = product_info.get('price', 0)
                            total_price = float(price) * item.box_quantities.get(box_number, 0) if price else 0
                            item.product_name = product_info.get('cn_name', item.product_name)
                        
                        # 累计总数和总金额
                        total_quantity += item.box_quantities.get(box_number, 0)
                        total_amount += total_price
                        total_weight += box.weight

                        # 设置单元格值
                        cell_data = [
                            (1, f"{code}U00000{box_number}" if item == box.items[0] else ""),  # FBA号,只在第一行显示
                            (2, box_number if item == box.items[0] else ""),  # 箱号,只在第一行显示
                            (3, product_info.get('cn_name', '') if product_info else ''),  # 中文品名
                            (4, product_info.get('en_name', '') if product_info else ''),  # 英文品名
                            (5, price),  # 单价
                            (6, item.box_quantities.get(box_number, 0)),  # 数量
                            (7, total_price),  # 总价
                            (8, f"{product_info.get('material_cn', '')}/{product_info.get('material_en', '')}" if product_info else ''),  # 材质
                            (9, f"{product_info.get('usage_cn', '')}/{product_info.get('usage_en', '')}" if product_info else ''),  # 用途
                            (10, box.weight if item == box.items[0] else ""),  # 毛重,只在第一行显示
                            (11, box.length if item == box.items[0] else ""),  # 长,只在第一行显示
                            (12, box.width if item == box.items[0] else ""),  # 宽,只在第一行显示
                            (13, box.height if item == box.items[0] else ""),  # 高,只在第一行显示
                            (14, product_info.get('brand', '') if product_info else ''),  # 品牌
                            (15, product_info.get('hs_code', '') if product_info else '')  # HS编码
                        ]
                        
                        # 批量设置单元格值和样式
                        for column, value in cell_data:
                            self._set_cell_value(sheet, row_num, column, value, style_info)

                        sheet.row_dimensions[row_num].height = row_height
                        sheet.column_dimensions['O'].width = row_height
                        row_num += 1
                    
                    # 如果这个箱子有多个产品,需要合并单元格
                    if box_items_count > 1:
                        merge_columns = [1, 2, 10, 11, 12, 13]  # 需要合并的列
                        for col in merge_columns:
                            self.merge_cells_in_range(sheet, start_row, row_num-1, col, col)

                # 添加总计行
                total_row = row_num  # 直接使用当前行号，不再加1
                self._set_cell_value(sheet, total_row, 1, "总件数", style_info)
                self._set_cell_value(sheet, total_row, 2, len(box_data), style_info)
                self._set_cell_value(sheet, total_row, 6, total_quantity, style_info)
                self._set_cell_value(sheet, total_row, 7, total_amount, style_info)
                self._set_cell_value(sheet, total_row, 9, "总重", style_info)
                self._set_cell_value(sheet, total_row, 10, total_weight, style_info)
                
                def set_cell_value(sheet, row, column, value, font_size=12):
                    cell = sheet.cell(row=row, column=column)
                    cell.value = value
                    cell.font = Font(name='Arial', size=font_size,bold=True)
                    no_border = Border(left=Side(border_style=None),
                                right=Side(border_style=None),
                                top=Side(border_style=None),
                                bottom=Side(border_style=None))
                    cell.border = no_border
                    
                def add_border_to_range(sheet, start_row, end_row, start_col, end_col):
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    for row in range(start_row, end_row + 1):
                        for col in range(start_col, end_col + 1):
                            cell = sheet.cell(row=row, column=col)
                            cell.border = border

                    # 为数据区域添加边框（从表头到合计行）
                add_border_to_range(sheet,total_row, total_row, 1, 15)

                # 使用辅助函数设置单元格
                made_in_row = total_row + 2
                current_date = datetime.now().strftime("%Y.%m.%d")
                set_cell_value(sheet, made_in_row, 11, "DATE")
                set_cell_value(sheet, made_in_row+1, 11, f"签字日期:{current_date}")
                set_cell_value(sheet, made_in_row, 1, "Made in China")

                # 设置行高
                # for row in range(13, row_num):
                #     sheet.row_dimensions[row].height = 
                self.merge_cells_in_range(sheet, 4, 4, 1, 3)
                self.merge_cells_in_range(sheet, 7, 11, 1, 3)
                self.merge_cells_in_range(sheet, 7, 11, 4, 15)

            except Exception as e:
                print(f"填充模板时发生错误: {str(e)}")
                raise

    @template_handler("林道")
    def _fill_lindao_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """
        填充林道模板
        :param wb: Excel工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                sheet = wb['模板']  # 获取模板工作表
                self.unmerge_cells_in_range(sheet, 3, 3, 2, 4)
                self.unmerge_cells_in_range(sheet, 4, 4, 2, 4)
                self.unmerge_cells_in_range(sheet, 5, 5, 2, 4)
                self.unmerge_cells_in_range(sheet, 6, 6, 2, 4)
                self.unmerge_cells_in_range(sheet, 9, 9, 2, 4)
                self.unmerge_cells_in_range(sheet, 11, 11, 2, 4)
                self.unmerge_cells_in_range(sheet, 12, 12, 2, 4)
                self.unmerge_cells_in_range(sheet, 15, 15, 2, 4)

                self.unmerge_cells_in_range(sheet, 13, 13, 6, 8)
                self.unmerge_cells_in_range(sheet, 14, 14, 6, 8)
                self.unmerge_cells_in_range(sheet, 15, 15, 6, 8)
                print("开始写入林道模版信息")

                box_Reference_id = '' 

                # 定义样式信息
                style_info = {
                    'font': Font(name='Arial', size=10),
                    'border': Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin')),
                    'alignment': Alignment(horizontal='center', vertical='center')
                }

                # 在第一行B列填充编码
                if code:
                    cell = sheet.cell(row=1, column=2)  # B列是第2列
                    cell.value = code
                    cell.font = Font(name='Arial', size=9)

                    cell = sheet.cell(row=14, column=6)  # B列是第2列
                    cell.value = code
                    cell.font = Font(name='Arial', size=9)
                    cell = sheet.cell(row=13, column=6)  # B列是第2列
                    cell.value = code
                    cell.font = Font(name='Arial', size=9)



                # 如果有地址信息，填充到相应的单元格
                if address_info:
                    address_info_detail = address_info['address_info']

                    if address_info_detail['amazonReferenceId']:
                        box_Reference_id =address_info_detail['amazonReferenceId']

                        cell = sheet.cell(row=15, column=6)  # B列是第2列
                        cell.value = box_Reference_id
                        cell.font = Font(name='Arial', size=9)

                    try:
                        # 填充收件人信息，这里收件人和
                        
                        # 填充地址信息
                        address_parts = []
                        if 'name' in address_info_detail:
                            cell = sheet.cell(row=4, column=2)  # B2单元格
                            cell.value = address_info_detail['name']

                            cell = sheet.cell(row=3, column=2)  # B2单元格
                            cell.value = address_info_detail['name']

                            cell = sheet.cell(row=5, column=2)  # B2单元格
                            cell.value = address_info_detail['name']
                            address_parts.append(address_info_detail['name'])
                        if 'addressLine1' in address_info_detail:
                            address_parts.append(address_info_detail['addressLine1'])
                        if 'city' in address_info_detail:
                            address_parts.append(address_info_detail['city'])
                        if 'stateOrProvinceCode' in address_info_detail:
                            address_parts.append(address_info_detail['stateOrProvinceCode'])
                        if 'postalCode' in address_info_detail:
                            address_parts.append(address_info_detail['postalCode'])
                        if 'countryCode' in address_info_detail:
                            address_parts.append(address_info_detail['countryCode'])

                        # 填充地址信息
                        if 'addressLine1' in address_info_detail:
                            cell = sheet.cell(row=6, column=2)  # B3单元格
                            cell.value = address_info_detail['addressLine1']

                        # 城市
                        if 'city' in address_info_detail:
                            cell = sheet.cell(row=9, column=2)  # B4单元格
                            cell.value = address_info_detail['city']

                        #邮政编码
                        if 'postalCode' in address_info_detail:
                            cell = sheet.cell(row=11, column=2)  # B6单元格
                            cell.value = address_info_detail['postalCode']

                        #国家代码
                        if 'countryCode' in address_info_detail:
                            cell = sheet.cell(row=12, column=2)  # B7单元格
                            cell.value = address_info_detail['countryCode']

                        if address_parts:
                            cell = sheet.cell(row=7, column=2)  # B3单元格
                            cell.value = ', '.join(address_parts)
                    except Exception as e:
                        print(f"填充地址信息时发生错误: {str(e)}")

                try:
                    total_boxes = len(box_data.keys())
                    cell = sheet.cell(row=15, column=2)  # 在第7行B列填充箱数
                    cell.value = str(total_boxes)
                    cell.font = Font(name='Arial', size=9)
                except Exception as e:
                    print(f"填充箱数时发生错误: {str(e)}")

                # 检查所有产品的电磁属性
                has_electric = False
                has_magnetic = False
                for box in box_data.values():
                    for item in box.items:
                        product_info = self._get_product_info(item.msku, db)
                        if product_info:
                            if product_info.get('electrified', '') == '是':
                                has_electric = True
                            if product_info.get('magnetic', '') == '是':
                                has_magnetic = True
                            if has_electric and has_magnetic:
                                break
                    if has_electric and has_magnetic:
                        break

                # 在表格顶部添加电磁属性标记
                if has_electric:
                    cell = sheet.cell(row=1, column=6)  # F列第1行
                    cell.value = "是"
                    cell.font = Font(name='Arial', size=9)
                
                if has_magnetic:
                    cell = sheet.cell(row=2, column=6)  # F列第2行
                    cell.value = "是"
                    cell.font = Font(name='Arial', size=9)

                # 填充数据
                row_num = 17  # 从第18行开始填充
                index = 1    # 添加序号计数器，从1开始
                row_height = sheet.row_dimensions[17].height

                # 遍历每个箱子

                sorted_boxes = sorted(box_data.items(), key=lambda x: int(x[0]))
                
                # 遍历排序后的箱子
                for box_number, box in sorted_boxes:
                    print(f"处理箱子 {box_number}")

                    # 遍历箱子中的每个产品
                    for item in box.items:
                        # 从数据库获取产品信息
                        product_info = self._get_product_info(item.msku, db)

                        Reference_id = None  # 初始化为None
                        if box_Reference_id:
                            Reference_id = box_Reference_id
                        # 处理产品信息为None的情况
                        if product_info is None:
                            print(f"警告: 未找到产品 {item.msku} 的信息")
                            price = 0
                            total_price = 0
                        else:
                            price = product_info.get('price', 0)
                            total_price = float(price) * item.box_quantities.get(box_number, 0) if price else 0
                            item.product_name = product_info.get('cn_name', item.product_name)

                        
                        # 设置单元格值和样式ç
                        cell_data = [
                            (1, box_number),                    # 货箱编号 (A列)
                            (2, code),
                            (3,Reference_id if Reference_id is not None else ""), 
                            (4, box.weight if box.weight is not None else ""),  # 重量 (B列)
                            (5, box.length if box.length is not None else ""),  # 长度 (Q列)
                            (6, box.width if box.width is not None else ""),    # 宽度 (R列)
                            (7, box.height if box.height is not None else ""),   # 高度 (S列)
                            (8,product_info.get('en_name', '') if product_info else ''),  # 链接 (D列)
                            (9, product_info.get('cn_name', '') if product_info else ''),  # 链接 (D列)
                            (10, product_info.get('price', '') if product_info else ''),   # 仅在总价格大于0时填入
                            (11,"美元"),  
                            (12, item.box_quantities.get(box_number, 0)),  # 数量 (F列)
                            (13, str(product_info.get('material_en', '')+'/'+product_info.get('material_cn', '')) if product_info else ''),  # 材料 (D列) 
                            (14, product_info.get('hs_code', '') if product_info else ''),  # HS编码 (G列)
                            (15, str(product_info.get('usage_en', '')+'/'+product_info.get('usage_cn', '' ))if product_info else ''),    # 用途 (H列)
                            (16, product_info.get('brand', '') if product_info else ''),    # 品牌 (I列)
                            (17, product_info.get('model', '') if product_info else ''),   # 型号 (J列)
                            # (12, product_info.get('link', '') if product_info else ''),
                            (18, ''),  # 图片列 (N列)
                            # (15, total_price if total_price > 0 else ""),  # 仅在总价格大于0时填入
                            # (24,item.SKU)
                        ]

                        # 批量设置单元格值和样式
                        for column, value in cell_data:
                            self._set_cell_value(sheet, row_num, column, value, style_info)

                        sheet.row_dimensions[row_num].height = row_height
                        sheet.column_dimensions['R'].width = row_height
                        # 插入产品图片
                        if item.msku and hasattr(self, 'image_folder'):
                            try:
                                image_cell = f"R{row_num}" 
                                # self.insert_product_image(sheet, image_cell, item.msku, self.image_folder)
                                self.insert_original_product_image(sheet, image_cell, item.msku, self.image_folder)
                            except Exception as e:
                                print(f"插入图片时发生错误: {str(e)}")

                        row_num += 1

                self.merge_cells_in_range(sheet, 3, 3, 2, 4)
                self.merge_cells_in_range(sheet, 4, 4, 2, 4)
                self.merge_cells_in_range(sheet, 5, 5, 2, 4)
                self.merge_cells_in_range(sheet, 6, 6, 2, 4)
                self.merge_cells_in_range(sheet, 9, 9, 2, 4)
                self.merge_cells_in_range(sheet, 11, 11, 2, 4)
                self.merge_cells_in_range(sheet, 12, 12, 2, 4)
                self.merge_cells_in_range(sheet, 15, 15, 2, 4)
                self.merge_cells_in_range(sheet, 13, 13, 6, 8)
                self.merge_cells_in_range(sheet, 14, 14, 6, 8)
                self.merge_cells_in_range(sheet, 15, 15, 6, 8)
            except Exception as e:
                print(f"填充模板时发生错误: {str(e)}")
                raise

    @template_handler("林道UPS")
    def _fill_lindaoUPS_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """
        填充林道UPS模板
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        try:
            sheet = wb['发票']
            
            # 解除单元格合并
            ranges_to_unmerge = ['A22:D22', 'A23:D23', 'A24:D24', 'A25:D25',
                               'A26:D26', 'A27:D27']
            for range_str in ranges_to_unmerge:
                try:
                    start_row, end_row, start_col, end_col = self._parse_range(range_str)
                    self.unmerge_cells_in_range(sheet, start_row, end_row, start_col, end_col)
                except Exception as e:
                    print(f"解除单元格合并时出错 {range_str}: {str(e)}")

            # 保存特定行的高度
            row_23_height = sheet.row_dimensions[23].height if 23 in sheet.row_dimensions else 15
            row_28_height = sheet.row_dimensions[28].height if 28 in sheet.row_dimensions else 15

            # 设置行高
            default_height = sheet.row_dimensions[60].height if 60 in sheet.row_dimensions else 15
            for r in range(19, 22):
                sheet.row_dimensions[r].height = default_height

            # 记录E24单元格的格式
            cell_e24 = sheet.cell(row=24, column=5)
            cell_font_e24 = Font(
                name=cell_e24.font.name if cell_e24.font.name else 'Arial',
                size=cell_e24.font.size if cell_e24.font.size else 11,
                bold=cell_e24.font.bold,
                italic=cell_e24.font.italic,
                vertAlign=cell_e24.font.vertAlign,
                color=cell_e24.font.color
            )
            cell_alignment_e24 = Alignment(
                horizontal=cell_e24.alignment.horizontal if cell_e24.alignment.horizontal else 'center',
                vertical=cell_e24.alignment.vertical if cell_e24.alignment.vertical else 'center',
                text_rotation=cell_e24.alignment.text_rotation,
                wrap_text=cell_e24.alignment.wrap_text,
                shrink_to_fit=cell_e24.alignment.shrink_to_fit,
                indent=cell_e24.alignment.indent
            )

            # 删除指定行
            sheet.delete_rows(22, 29)

            # 记录格式信息
            row_height = sheet.row_dimensions[19].height if 19 in sheet.row_dimensions else 15
            cell_page = sheet.cell(row=19, column=1)
            cell_border = Border(
                left=cell_page.border.left if cell_page.border.left else Side(style='thin'),
                right=cell_page.border.right if cell_page.border.right else Side(style='thin'),
                top=cell_page.border.top if cell_page.border.top else Side(style='thin'),
                bottom=cell_page.border.bottom if cell_page.border.bottom else Side(style='thin')
            )
            cell_font = Font(
                name=cell_page.font.name if cell_page.font.name else 'Arial',
                size=cell_page.font.size if cell_page.font.size else 11,
                bold=cell_page.font.bold,
                italic=cell_page.font.italic,
                vertAlign=cell_page.font.vertAlign
            )
            cell_alignment = Alignment(horizontal='center', vertical='center')
            # 填充数据
            num_row = 19
            for box_number, box in box_data.items():
                if not box.items:
                    continue
                for product_info in box.items:
                    # 获取产品信息
                    db_product_info = self._get_product_info(product_info.msku, db)
                    if not db_product_info:
                        continue

                    # 构建产品名称和获取数量、价格
                    name = f"{db_product_info.get('en_name', '')}({db_product_info.get('cn_name', '')})"
                    quantity = product_info.box_quantities.get(box_number, 0)
                    price = db_product_info.get('price', 0)

                    # 设置单元格值
                    cell_values = [
                        (1, name), 
                        (2, quantity), 
                        (3, price), 
                        (4, quantity * price),
                        ('CN', 5)
                    ]
                    
                    for value, col in cell_values:
                        cell = sheet.cell(row=num_row, column=col, value=value)
                        cell.font = cell_font
                        cell.alignment = cell_alignment
                        cell.border = cell_border
                    num_row += 1

            # 设置行高和边框
            for row in range(19, num_row + 8):
                sheet.row_dimensions[row].height = row_height
                for col in range(1, 6):
                    cell = sheet.cell(row=row, column=col)
                    if not cell.border:
                        cell.border = cell_border

            # 添加底部文本
            declarations = [
                ('THESE COMMODITIES ARE LICENSED FOR THE UNTIMATE DESTINATION SHOWN.', cell_font),
                ('以上商品已有到最终目的地的许可。', cell_font),
                ('', None),
                ('I DECLARE ALL THE INFORMATION CONTAINED IN THIS INVOICE LIST TO BE TRUE AND CORRECT.',
                 Font(name='Arial', size=9, color='000080')),
                ('以上申报均属实。', Font(name='宋体', size=11, color='FF0000', bold=True)),
                ('', None),
                ('SIGNATURE OF SHIPPER/EXPORTER(TYPE NAME TITLE AND SIGN):    ',
                 Font(name='Arial', size=9, color='000080', bold=True)),
                ('寄件人/出口商签名(正楷和职位)', Font(name='宋体', size=9, color='000080', bold=True))
            ]

            for i, (text, font) in enumerate(declarations):
                if text:
                    cell = sheet.cell(row=num_row + i, column=1, value=text)
                    if font:
                        cell.font = font
                    cell.alignment = cell_alignment
                    if i == 1:
                        sheet.row_dimensions[num_row + i].height = row_23_height
                    elif i == 6:
                        sheet.row_dimensions[num_row + i].height = row_28_height

            # 合并单元格
            for row in range(num_row, num_row + 8):
                try:
                    self.merge_cells_in_range(sheet, row, row, 1, 4)
                except Exception as e:
                    print(f"合并单元格时出错 row {row}: {str(e)}")

            # 设置右侧文本
            right_text = [
                (num_row, 5, 'CHECK ONE', cell_font),
                (num_row + 1, 5, '□ F.O.B', cell_font),
                (num_row + 2, 5, '', cell_font_e24),
                (num_row + 6, 4, 'DATE:', cell_font),
                (num_row + 7, 4, '日期', cell_font)
            ]

            for row, col, text, font in right_text:
                cell = sheet.cell(row=row, column=col, value=text)
                cell.font = font
                cell.alignment = (cell_alignment_e24 if col == 5 and row == num_row + 2 
                                else cell_alignment)

        except Exception as e:
            print(f"填充林道UPS模板时发生错误: {str(e)}")
            raise

    @template_handler("递信")
    def _fill_dixing_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """
        填充递信模板
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                
                sheet = wb['FBA对应贴标资料']  # 获取模板工作表
                print("开始写入递信模版信息")
                current_date = datetime.now().strftime("%Y.%m.%d")
                cell = sheet.cell(row=1, column=4)
                cell.value = current_date
                cell.font = Font(name='Arial', size=12)

                sheet_shipment_id = shipment_id if shipment_id else ''
                # 记录第3行的格式信息
                row_height = sheet.row_dimensions[3].height

                # 查找并取消合并单元格
                merged_cells = sheet.merged_cells
                cells_to_unmerge = []
                for merged_cell in merged_cells:
                    if merged_cell.min_row >= 3:
                        cells_to_unmerge.append(merged_cell)

                for cell_range in cells_to_unmerge:
                    sheet.unmerge_cells(str(cell_range))

                # 保存最后一行的高度
                last_height = sheet.row_dimensions[21].height

                # 删除原来的内容
                sheet.delete_rows(3, 20)

                address_parts = []
                adress = ''
                # 如果有地址信息，填充到相应的单元格
                if address_info:
                    address_info_detail = address_info['address_info']
                    try:
                        # 填充收件人信息，这里收件人和
                        
                        # 填充地址信息
                       
                        if 'name' in address_info_detail:
                            address_parts.append(address_info_detail['name'])
                        if 'addressLine1' in address_info_detail:
                            address_parts.append(address_info_detail['addressLine1'])
                        if 'city' in address_info_detail:
                            address_parts.append(address_info_detail['city'])
                        if 'stateOrProvinceCode' in address_info_detail:
                            address_parts.append(address_info_detail['stateOrProvinceCode'])
                        if 'postalCode' in address_info_detail:
                            address_parts.append(address_info_detail['postalCode'])
                        if 'countryCode' in address_info_detail:
                            address_parts.append(address_info_detail['countryCode'])

                        if address_parts:
                            # cell = sheet.cell(row=3, column=14)  # B3单元格
                            # adress = ', '.join(address_parts)
                            adress = ', \n'.join(address_parts)

                    except Exception as e:
                        print(f"填充地址信息时发生错误: {str(e)}")
                

                # 设置行高
                for r in range(3, 21):
                    sheet.row_dimensions[r].height = sheet.row_dimensions[25].height

                # 初始化统计数据
                total_weight = 0
                total_quantity = 0
                row_num = 3
                ticket = str(code)+"U00000"

                # 设置单元格对齐方式
                center_alignment = Alignment(horizontal='center', vertical='center')

                # 遍历每个箱子

                for box_number, box in sorted(box_data.items(), key=lambda x: int(x[0])):
                    # 计算箱子中的产品数量
                    box_products = box.items
                    print(box_products)
                    is_mixed = len(box_products) >= 2
                    # identifier = f"{len(box_data)}-{box_number}{'(混装)' if is_mixed else ''}"
                    identifier = f"{sheet_shipment_id}0{box_number}"

                    merge_start_row = row_num
                    
                    # 计算该箱子的总数量
                    box_total_quantity = sum(
                        getattr(product_info, 'box_quantities', {}).get(box_number, 0)
                        for product_info in box_products
                    )
                    print(f"Box {box_number} total quantity: {box_total_quantity}")

                    # 遍历箱子中的每个产品
                    for index, product_info in enumerate(box_products):
                        if hasattr(product_info, 'msku'):
                            db_product_info = self._get_product_info(product_info.msku, db)
                            if db_product_info:
                                # 更新产品信息
                                for key, value in db_product_info.items():
                                    setattr(product_info, key, value)

                        # 设置单元格值
                        cell_data = [
                            (1, identifier),  # 标识符
                            (2, f"{getattr(product_info, 'cn_name', '')}\n{getattr(product_info, 'en_name', '')}"),  # 品名
                            (3, f"{getattr(product_info, 'material_cn', '')}\n{getattr(product_info, 'material_en', '')}"),  # 材质
                            (4, f"{getattr(product_info, 'usage_en', '')}, {getattr(product_info, 'usage_cn', '')}"),  # 用途
                            (5, box_number),  # 箱号
                            (8, box_total_quantity), 
                            (6, getattr(box, 'weight', '')),  # 重量
                            (7, f"{getattr(box, 'length', '')}*{getattr(box, 'width', '')}*{getattr(box, 'height', '')}"),  # 尺寸
                        ]

                        # 处理数量信息
                        quantity = getattr(product_info, 'box_quantities', {}).get(box_number, 0)
                        original_value = getattr(product_info, 'box_original_values', {}).get(box_number, str(quantity))
                        
                        # print("数量信息:")
                        # print(f"  - 数字形式: {quantity}")
                        # print(f"  - 原始格式: {original_value}")
                        
                        total_quantity = total_quantity + quantity

                        # 只在第一行显示箱子的总数量
                        if quantity:
                            if ' ' in original_value:
                                prefix, number = original_value.split(' ', 1)
                                cell_data.extend([
                                    # (8, number),  # 数量
                                    (9, number),  # 数量（重复）
                                    (10, f"{prefix} {str(quantity)}"),  # 前缀（如 A1）
                                ])
                            else:
                                cell_data.extend([
                                    (9, str(box_total_quantity)),  # 数量（重复）
                                    (10, str(box_total_quantity)),  # 没有前缀
                                ])
                        else:
                            cell_data.extend([
                                (9, ''),  # 数量（重复）
                                (10, ''),  # 没有前缀
                            ])

                        # 设置运单号
                        cell_data.append((13, f"{ticket}{box_number}"))  # 运单号
                        cell_data.append((14, adress))  # 地址信息

                        # 设置单元格值和格式
                        for col, value in cell_data:
                            cell = sheet.cell(row=row_num, column=col, value=value)
                            cell.alignment = center_alignment

                        # 插入产品图片
                        if hasattr(product_info, 'msku') and hasattr(self, 'image_folder'):
                            try:
                                image_cell = f"L{row_num}"
                                sheet.row_dimensions[row_num].height = 95
                                # self.insert_product_image(sheet, image_cell, product_info.msku, self.image_folder)
                                self.insert_original_product_image(sheet, image_cell, product_info.msku, self.image_folder)
                            except Exception as e:
                                print(f"插入图片时发生错误: {str(e)}")

                        row_num += 1
                        
                    # 合并单元格
                    if row_num > merge_start_row:
                        merge_ranges = [
                            (merge_start_row, 1, row_num - 1, 1),  # 标识符列
                            (merge_start_row, 5, row_num - 1, 5),  # 箱号列
                            (merge_start_row, 6, row_num - 1, 6),  # 重量列
                            (merge_start_row, 7, row_num - 1, 7),  # 尺寸列
                            (merge_start_row, 8, row_num - 1, 8),  # 数量列
                            (merge_start_row, 13, row_num - 1, 13),  # 运单号列
                           
                        ]
                        for start_row, start_col, end_row, end_col in merge_ranges:
                            sheet.merge_cells(
                                start_row=start_row,
                                start_column=start_col,
                                end_row=end_row,
                                end_column=end_col
                            )
                    total_weight += box.weight

                # 合并最后一列
                sheet.merge_cells(start_row=3, start_column=14, end_row=row_num, end_column=14)

                # 删除多余的行
                if row_num < sheet.max_row:
                    sheet.delete_rows(row_num + 1, sheet.max_row - row_num)

                # 设置边框
                cell_border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )

                # 设置所有单元格的边框和行高
                for row in range(3, row_num + 1):
                    sheet.row_dimensions[row].height = row_height
                    sheet.column_dimensions['L'].width = row_height
                    for col in range(1, 15):
                        cell = sheet.cell(row=row, column=col)
                        cell.border = cell_border

                # 设置最后一行的汇总信息
                sheet.row_dimensions[row_num].height = last_height

                # 设置汇总行的字体和样式
                summary_data = [
                    (1, '汇总', Font(bold=True, name='宋体', size=12)),
                    (5, len(box_data), Font(bold=True, name='微软雅黑', size=11)),
                    (6, total_weight, Font(bold=True, name='微软雅黑', size=9)),
                    (8, total_quantity, Font(bold=True, name='微软雅黑', size=12)),
                    (9, total_quantity, Font(bold=True, name='微软雅黑', size=11))
                ]

                for col, value, font in summary_data:
                    cell = sheet.cell(row=row_num, column=col, value=value)
                    cell.font = font
                    cell.alignment = center_alignment

                print("递信模板填充完成")

            except Exception as e:
                print(f"填充模板时发生错误: {str(e)}")
                traceback.print_exc()
                raise
    
    @template_handler("德邦美森限时达")
    def _fill_dbmsxsd_template(self, wb, box_data,code=None, address_info=None, shipment_id=None):
        """
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                sheet = wb['FBA专线出货资料模板']  # 获取模板工作表
                total_quantity = 0
                total_amount = 0
                total_weight = 0
                print("开始写入模版信息")

                # 定义样式信息
                style_info = {
                    'font': Font(name='Arial', size=10),
                    'border': Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin')),
                    'alignment': Alignment(horizontal='center', vertical='center')
                }

                #先拆分合并的单元格，用于写入
                self.unmerge_cells_in_range(sheet, 2, 2, 3, 5)
                self.unmerge_cells_in_range(sheet, 3, 3, 3, 5)
                self.unmerge_cells_in_range(sheet, 4, 4, 3, 5)
                # self.unmerge_cells_in_range(sheet, 4, 4, 7, 8)
                self.unmerge_cells_in_range(sheet, 5, 5, 3, 5)

                if code:
                    cell = sheet.cell(row=4, column=7)  # B列是第2列
                    cell.value = "FBA 号：" + str(code)
                    cell.font = Font(name='Arial', size=12,bold=True)

                # 如果有地址信息，填充到相应的单元格
                if address_info:
                    address_info_detail = address_info['address_info']
                    # if 'seller_info' in address_info:
                    #     box_Reference_id = address_info['seller_info']['amazonReferenceId']
                    try:
                        # 填充收件人信息
                       
                        # 填充地址信息
                        address_parts = []
                        if 'name' in address_info_detail:
                            cell = sheet.cell(row=3, column=3)  
                            cell.value = address_info_detail['name']
                            cell = sheet.cell(row=4, column=3)  
                            cell.value = address_info_detail['name']
                            address_parts.append(address_info_detail['name'])
                        if 'addressLine1' in address_info_detail:
                            address_parts.append(address_info_detail['addressLine1'])
                        if 'city' in address_info_detail:
                            address_parts.append(address_info_detail['city'])
                        if 'stateOrProvinceCode' in address_info_detail:
                            address_parts.append(address_info_detail['stateOrProvinceCode'])
                        if 'postalCode' in address_info_detail:
                            address_parts.append(address_info_detail['postalCode'])
                        if 'countryCode' in address_info_detail:
                            address_parts.append(address_info_detail['countryCode'])

                        if address_parts:
                            cell = sheet.cell(row=2, column=3)  
                            cell.value = ', '.join(address_parts)
                            
                            cell = sheet.cell(row=5, column=3)  
                            cell.value = ', '.join(address_parts)
  
                    except Exception as e:
                        print(f"填充地址信息时发生错误: {str(e)}")

                # 填充数据
                row_num = 9  
                index = 1    # 添加序号计数器，从1开始
                row_height = sheet.row_dimensions[9].height

                # 将box_data按箱号排序
                sorted_boxes = sorted(box_data.items(), key=lambda x: int(x[0]))
                
                # 遍历排序后的箱子
                for box_number, box in sorted_boxes:
                    print(f"处理箱子 {box_number}")
                    first_row_of_box = row_num  # 记录这个箱子的第一行
                    box_number_str = code + 'U00000' + str(box_number)

                    # 遍历箱子中的每个产品
                    for item in box.items:
                        # 从数据库获取产品信息
                        product_info = self._get_product_info(item.msku, db)
                        volume = box.length * box.width * box.height*0.000001
                        price = 0
                        total_price = 0

                        total_quantity += item.box_quantities.get(box_number, 0)
                        total_amount += total_price
                        total_weight += box.weight
    
                        if product_info is not None:
                            item.product_name = product_info.get('cn_name', item.product_name)
                            print(f"产品信息：{product_info}")
                        else:
                        # 处理未找到产品信息的情况
                            print(f"未找到产品信息，MSKU: {item.msku}")
                            item.product_name = "需要补数据"  # 可以设置一个默认值
                    
                        # 设置单元格值和样式
                        cell_data = [
                            # 基本信息
                            # 产品名称信息
                            (3, f"{product_info.get('en_name', '')} ({product_info.get('cn_name', '')})" if product_info else ''),    
                            (6,''),
                            (11, product_info.get('model', '') if product_info else ''),                   # 型号
                            # 产品材料和用途
                            (7,''),
                            (8, f"{product_info.get('material_en', '')} /{product_info.get('material_cn', '')}" if product_info else ''),            # 中文材料
                            (9, str(product_info.get('usage_en', '') + '/' +
                                   product_info.get('usage_cn', '')) if product_info else ''),            # 用途
                            (2, product_info.get('hs_code', '') if product_info else ''),                # HS编码
                            (5, item.box_quantities.get(box_number, 0)),         # 数量
                            (10,  product_info.get('electrified', '')if product_info else ''),            
                            (4, ''),                                            # 图片占位
                        ]
                        # 批量设置单元格值和样式
                        for column, value in cell_data:
                            self._set_cell_value(sheet, row_num, column, value, style_info)

                        sheet.row_dimensions[row_num].height = row_height
                        
                        # 插入产品图片
                        if item.msku and hasattr(self, 'image_folder'):
                            try:
                                image_cell = f"D{row_num}"  # 图片列（第14列）
                                # self.insert_product_image(sheet, image_cell, item.msku, self.image_folder)
                                self.insert_original_product_image(sheet, image_cell, item.msku, self.image_folder)
                            except Exception as e:
                                print(f"插入图片时发生错误: {str(e)}")

                        row_num += 1
                        
                    box_info_data = [
                        (12, box_number_str),  
                        (13, box.weight if box.weight is not None else ""),     # 重量
                        (14, box.weight if box.weight is not None else ""),     # 重量
                        (15, volume if volume is not None else "")          # 体积
                    ]

                    # 设置箱子信息
                    for column, value in box_info_data:
                        cell = sheet.cell(row=first_row_of_box, column=column, value=value)
                        cell.font = style_info['font']
                        cell.border = style_info['border']
                        cell.alignment = style_info['alignment']

                    # 使用箱子中的产品数量来确定合并范围
                    if len(box.items) > 1:  # 只有当箱子中有多个产品时才合并
                        for column, _ in box_info_data:
                            sheet.merge_cells(
                                start_row=first_row_of_box,
                                start_column=column,
                                end_row=first_row_of_box + len(box.items) - 1,
                                end_column=column
                            )
                               # 添加总计行

                total_row = row_num  # 直接使用当前行号，不再加1
                self._set_cell_value(sheet, total_row, 2, "TOTAL", style_info)
                self._set_cell_value(sheet, total_row, 12, len(box_data), style_info)
                self._set_cell_value(sheet, total_row, 5, total_quantity, style_info)
                # self._set_cell_value(sheet, total_row, 7, total_amount, style_info)
             
                self._set_cell_value(sheet, total_row, 13, total_weight, style_info)
                self._set_cell_value(sheet,total_row,14,total_weight,style_info)
                
                self.merge_cells_in_range(sheet, 2, 2, 3, 5)
                self.merge_cells_in_range(sheet, 3, 3, 3, 5)
                self.merge_cells_in_range(sheet, 4, 4, 3, 5)
                self.merge_cells_in_range(sheet, 4, 4, 7, 8)
                self.merge_cells_in_range(sheet, 5, 5, 3, 5)
                
                thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(border_style='thin'))

                for row_index in range(total_row-1, total_row+1):  # 行索引从1到3（对应A1:C3中的1到3行）
                    for col_index in range(2, 16):  # 列索引从1到3（对应A、B、C三列）
                        cell = sheet.cell(row=row_index, column=col_index)
                        cell.border = thin_border

            except Exception as e:
                print(f"填充模板时发生错误: {str(e)}")
                raise

    @template_handler("德邦空派")
    def _fill_dbkp_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """
        填充德邦空派模板
        :param wb: 工作簿对象
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        """
        with self.db_connector as db:
            try:
                sheet = wb['箱单发票']  # 获取模板工作表
                
                # 获取基础单元格样式
                base_cell = sheet.cell(row=9, column=3)
                base_border = Border(
                    left=base_cell.border.left,
                    right=base_cell.border.right,
                    top=base_cell.border.top,
                    bottom=base_cell.border.bottom
                )
                base_font = Font(
                    name=base_cell.font.name,
                    size=base_cell.font.size,
                    bold=base_cell.font.bold,
                    italic=base_cell.font.italic,
                    vertAlign=base_cell.font.vertAlign
                )
                base_alignment = Alignment(horizontal='center', vertical='center')
                style_info = {'font': base_font, 'border': base_border, 'alignment': base_alignment}
                
                # 保存行高信息
                # row_height = sheet.row_dimensions[9].height
                row_height = 60
                row_height_low = 30
                
                # 计算总数据
                total_length = 0
                total_weight = 0
                for box in box_data.values():
                    total_length += len(box.items)
                    total_weight += box.weight if hasattr(box, 'weight') else 0
                
                # 插入所需行数
                if total_length > 5:
                    sheet.insert_rows(9, total_length - 5)
                
                self.unmerge_cells_in_range(sheet, 4, 4, 3, 7)
                self.unmerge_cells_in_range(sheet, 5, 5, 3, 7)
                self.unmerge_cells_in_range(sheet, 5, 5, 12, 18)

                # 解除数据区域内的合并单元格
                print("解除数据区域内的合并单元格...")
                try:
                    data_ranges = []
                    for merged_range in sheet.merged_cells.ranges:
                        range_str = str(merged_range)
                        start_row, end_row, _, _ = self._parse_range(range_str)
                        if start_row >= 9:  # 只解除数据区域的合并单元格
                            data_ranges.append(range_str)
                    
                    for range_str in data_ranges:
                        sheet.unmerge_cells(range_str)
                        print(f"解除合并单元格: {range_str}")
                except Exception as e:
                    print(f"解除合并单元格时发生错误: {str(e)}")
                    traceback.print_exc()
                
                if code:
                    cell = sheet.cell(row=4, column=3)  # B列是第2列
                    cell.value = code
                    cell.font = Font(name='Arial', size=9)

                # 如果有地址信息，填充到相应的单元格
                if address_info:
                    address_info_detail = address_info['address_info']
                    try:
                        address_parts = []
                        if 'name' in address_info_detail:
                            cell = sheet.cell(row=5, column=3)  # B2单元格
                            cell.value = address_info_detail['name']
                            address_parts.append(address_info_detail['name'])
                        if 'addressLine1' in address_info_detail:
                            address_parts.append(address_info_detail['addressLine1'])
                        if 'city' in address_info_detail:
                            address_parts.append(address_info_detail['city'])
                        if 'stateOrProvinceCode' in address_info_detail:
                            address_parts.append(address_info_detail['stateOrProvinceCode'])
                        if 'postalCode' in address_info_detail:
                            address_parts.append(address_info_detail['postalCode'])
                        if 'countryCode' in address_info_detail:
                            address_parts.append(address_info_detail['countryCode'])

                        if address_parts:
                            cell = sheet.cell(row=5, column=12)  # B3单元格
                            cell.value = ', '.join(address_parts)
                    except Exception as e:
                        print(f"填充地址信息时发生错误: {str(e)}")

                # 填充数据
                row_num = 9
                data_rows = []  # 存储所有产品数据行
                sheet.delete_rows(9)
                
                # 第一步：收集所有产品数据
                for box_number, box in sorted(box_data.items(), key=lambda x: int(x[0])):
                    box_start_row = row_num
                    total_quantity = 0
                    total_price = 0
                    
                    for item in box.items:
                        product_info = self._get_product_info(item.msku, db)
                        if not product_info:
                            continue
                            
                        # # 序号
                        # self._set_cell_value(sheet, row_num, 1, row_num - 8, style_info)
                        # FBA号
                        fba_number = f"FBA176FB5SRR200000{box_number}"
                        self._set_cell_value(sheet, row_num, 2, fba_number, style_info)
                        # 箱号
                        self._set_cell_value(sheet, row_num, 3, box_number, style_info)
                        # 产品名称
                        name = f"{product_info.get('en_name', '')}({product_info.get('cn_name', '')})"
                        self._set_cell_value(sheet, row_num, 4, name, style_info)
                        # HS编码
                        self._set_cell_value(sheet, row_num, 5, product_info.get('hs_code', ''), style_info)
                        # 数量
                        quantity = item.quantity if hasattr(item, 'quantity') else 0
                        self._set_cell_value(sheet, row_num, 6, quantity, style_info)

                        self._set_cell_value(sheet,row_num,16,'',style_info)
                        
                        # 单价
                        price_str = product_info.get('price', '')
                        price = float(price_str) if price_str else 0
                        self._set_cell_value(sheet, row_num, 7, f"${price}", style_info)
                        
                        # 总价
                        total = round(float(quantity) * price, 2)
                        self._set_cell_value(sheet, row_num, 8, f"${total}", style_info)
                        
                        # 重量相关
                        box_weight = box.weight if hasattr(box, 'weight') else 0
                        for col in range(9, 12):
                            self._set_cell_value(sheet, row_num, col, box_weight, style_info)
                        
                        # 箱子尺寸
                        if hasattr(box, 'length'):
                            self._set_cell_value(sheet, row_num, 12, box.length, style_info)
                            self._set_cell_value(sheet, row_num, 13, box.width, style_info)
                            self._set_cell_value(sheet, row_num, 14, box.height, style_info)
                            volume = box.length * box.width * box.height * 0.000001
                            self._set_cell_value(sheet, row_num, 15, volume, style_info)
                        
                        # 磁性
                        self._set_cell_value(sheet, row_num, 17, product_info.get('magnetic', ''), style_info)

                        self._set_cell_value(sheet, row_num, 18, product_info.get('link', ''), style_info)
                        
                        # 插入产品图片
                        # self.insert_product_image(sheet, f'P{row_num}', item.msku, self.image_folder)
                        self.insert_original_product_image(sheet, f'P{row_num}', item.msku, self.image_folder)
                        
                        total_quantity += quantity
                        total_price += total
                        
                        # 保存产品信息用于后续创建申报要素表格
                        data_rows.append({
                            'product_info': product_info,
                            'row': row_num
                        })
                        
                        row_num += 1
                    
                    # 合并相同箱号的单元格
                    if row_num - box_start_row > 1:
                        for col in [1, 2, 12, 13, 14, 15]:
                            try:
                                merge_range = f"{get_column_letter(col)}{box_start_row}:{get_column_letter(col)}{row_num-1}"
                                sheet.merge_cells(merge_range)
                            except Exception as e:
                                print(f"合并单元格失败 {merge_range}: {str(e)}")

                # 设置原始数据区域的行高
                for row in range(9, row_num):
                    sheet.row_dimensions[row].height = row_height
                    sheet.column_dimensions['P'].width = row_height
                
                # 空一行开始添加申报要素表格
                row_num += 2
                
                # 第二步：为每个产品创建申报要素表格
                for data_row in data_rows:
                    product_info = data_row['product_info']
                    
                    # 创建申报要素表格
                    table_height = self._create_declaration_table(sheet, row_num+1, product_info)
                    
                    # 设置表格区域的行高，暂时先不用
                    for i in range(row_num, row_num + table_height):
                        sheet.row_dimensions[i].height = row_height_low
                    
                    # 更新行号（表格高度 + 1行间距）
                    row_num += table_height + 1
                
                # 合并底部单元格
                try:
                    sheet.merge_cells(f"F{row_num+2}:M{row_num+8}")
                    sheet.merge_cells(f"B{row_num+2}:D{row_num+2}")
                    sheet.merge_cells(f"B{row_num+10}:D{row_num+10}")
                    sheet.merge_cells(f"B{row_num+18}:D{row_num+18}")
                    
                    for i in range(3, 9):
                        sheet.merge_cells(f"C{row_num+i}:D{row_num+i}")
                    
                    for i in range(11, 17):
                        sheet.merge_cells(f"C{row_num+i}:D{row_num+i}")

                    for i in range(19, 25):
                        sheet.merge_cells(f"C{row_num+i}:D{row_num+i}")

                except Exception as e:
                    print(f"合并底部单元格时发生错误: {str(e)}")

            except Exception as e:
                print(f"填充德邦空派模板时发生错误: {str(e)}")
                traceback.print_exc()
                raise ProcessingError(f"填充德邦空派模板失败: {str(e)}")

    def _fill_default_template(self, wb, box_data, code=None, address_info=None, shipment_id=None):
        """默认的模板处理方法"""
        raise ProcessingError("未找到匹配的模板处理方法，请确保模板文件名包含正确的关键字")

    def generate_invoice(self, template_path, box_data, code=None, address_info=None, shipment_id=None):
        """
        生成发票
        :param template_path: 模板文件路径
        :param box_data: 箱子数据
        :param code: 编码（可选）
        :param address_info: 地址信息（可选）
        :param shipment_id: Shipment ID（可选）
        :return: 生成的发票文件路径
        """
        try:
            print(f"开始处理模板文件: {template_path}")
            if not os.path.exists(template_path):
                raise ProcessingError(f"模板文件不存在: {template_path}")

            # 获取当前时间戳
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            
            # 检查产品的电磁属性
            has_electric = False
            has_magnetic = False
            for box in box_data.values():
                for item in box.items:
                    product_info = self._get_product_info(item.msku, self.db_connector)
                    if product_info:
                        if product_info.get('electrified', '') == '是':
                            has_electric = True
                        if product_info.get('magnetic', '') == '是':
                            has_magnetic = True
                        if has_electric and has_magnetic:
                            break
                if has_electric and has_magnetic:
                    break
            
            # 构建文件名后缀
            suffix = ""
            if has_electric:
                suffix += "_带电"
            if has_magnetic:
                suffix += "_带磁"

            # 确保输出目录存在
            if not os.path.exists(self.output_folder):
                os.makedirs(self.output_folder)

            # 构建输出文件路径
            if address_info:
                try:
                    shipment_name = address_info['address_info']['shipmentName']
                    # 使用正则表达式提取数据
                    time, logistics, number = self.extract_data(shipment_name)
                    
                    if logistics is not None:
                        logistics = logistics.replace('-', '')  # 去掉破折号
                    else:
                        raise ValueError('Logistics cannot be None')

                    # 获取code
                    code_suffix = f"-{code}" if code else ""

                    output_filename = f'百泰{code_suffix}-{time}-{logistics}票-{number}件-{address_info["seller_info"]["country_name"]}-发票装箱单.xlsx'
                    # 替换任何可能导致路径问题的字符
                    output_filename = "".join(c for c in output_filename if c not in r'<>:"/\|?*')
                    output_path = os.path.join(self.output_folder, output_filename)

                except (TypeError, ValueError) as e:
                    # 处理异常的情况
                    print(f"Error occurred: {e}. Using default filename.")
                    output_filename = f"{timestamp}{suffix}.xlsx"
                    output_path = os.path.join(self.output_folder, output_filename)
            else:
                output_filename = f"{timestamp}{suffix}.xlsx"
                output_path = os.path.join(self.output_folder, output_filename)

            # 使用openpyxl加载模板
            print(f"正在加载模板文件...")
            wb = load_workbook(template_path)
            print(f"成功加载模板文件，工作表: {wb.sheetnames}")

            # 获取对应的模板处理方法
            template_handler = self._get_template_handler(template_path)
            if template_handler is None:
                raise ProcessingError(f"未找到对应的模板处理方法: {template_path}")

            # 处理模板
            template_handler(wb, box_data, code, address_info, shipment_id)

            # 保存文件
            wb.save(output_path)
            print(f"发票已生成: {output_path}")

            return output_path

        except Exception as e:
            error_msg = f"生成发票时发生错误: {str(e)}"
            print(error_msg)
            traceback.print_exc()
            raise ProcessingError(error_msg)

    def _get_template_handler(self, template_path):
        """根据模板文件名选择对应的处理方法"""
        try:
            template_name = os.path.basename(template_path).lower()
            # print(f"正在查找模板处理器，模板路径: {template_path}")
            # print(f"模板文件名: {template_name}")
            # print(f"已注册的处理器: {self._template_handlers}")
            
            for keyword, handler in self._template_handlers.items():
                # print(f"检查关键字: {keyword}, 类型: {type(keyword)}")
                # print(f"模板名称: {template_name}, 类型: {type(template_name)}")
                keyword_lower = keyword.lower()
                if keyword_lower in template_name:
                    print(f"找到匹配的处理器: {handler.__name__}")
                    return handler.__get__(self, type(self))
            print(f"未找到匹配的处理器，可用的关键字: {list(self._template_handlers.keys())}")
            return self._fill_default_template
        except Exception as e:
            print(f"模板处理器匹配过程中出错: {str(e)}")
            return self._fill_default_template

    def _get_product_info(self, msku, db=None):
        """
        从MongoDB获取产品信息
        :param msku: 产品的MSKU
        :param db: 数据库连接（可选）
        :return: 包含产品信息的字典
        """
        try:
            if db is None:
                # 如果没有传入db连接，创建新的连接
                with self.db_connector as db:
                    return self._get_product_info(msku, db)
            
            # 使用传入的db连接
            collection = db['msku_info']
            product = collection.find_one({'msku': msku})
            
            if product:
                return {
                    'cn_name': product.get('productNameZh', ''),
                    'en_name': product.get('productNameEn', ''),
                    'en_usage': product.get('useEn', ''),
                    'ch_usage':product.get('useZh', ''),
                    'material_en': product.get('materialEn', ''),
                    'material_cn': product.get('materialZh', ''),
                    'hs_code': product.get('HS', ''),
                    'usage_en': product.get('useEn', ''),
                    'usage_cn': product.get('useZh', ''),
                    'brand': product.get('brand', ''),
                    'model': product.get('model', ''),
                    'link': product.get('productLink', ''),
                    'price':product.get('askprice', ''),
                    'electrified':product.get('electrified', ''),
                    'magnetic':product.get('magnetic', ''),
                    'weight':product.get('weight', ''),
                }
            return None
        except Exception as e:
            print(f"Error fetching product info for MSKU {msku}: {str(e)}")
            return None

    def _set_cell_value(self, sheet, row, column, value, style_info):
        """
        设置单元格的值和样式
        :param sheet: 工作表对象
        :param row: 行号
        :param column: 列号
        :param value: 单元格值
        :param style_info: 样式信息
        """
        cell = sheet.cell(row=row, column=column)
        cell.value = value
        cell.font = style_info['font']
        cell.border = style_info['border']
        cell.alignment = style_info['alignment']

    def insert_centered_image(self, worksheet, cell_address, image_path, fixed_width=None, fixed_height=None):
        """
        在指定的单元格中插入居中的图片
        :param worksheet: 工作表对象
        :param cell_address: 单元格地址（例如'A1'）
        :param image_path: 图片文件路径
        :param fixed_width: 固定宽度（可选）
        :param fixed_height: 固定高度（可选）
        :return: 是否成功插入图片
        """
        try:
            # 读取图片
            img = PILImage.open(image_path)
            
            # 获取单元格的宽度和高度（以像素为单位）
            column_width = worksheet.column_dimensions[cell_address[0]].width
            row_height = worksheet.row_dimensions[int(cell_address[1:])].height
            
            # 如果没有指定宽度和高度，使用单元格的大小
            if fixed_width is None:
                fixed_width = column_width * 7  # 转换为像素
            if fixed_height is None:
                fixed_height = row_height * 1.5  # 转换为像素
                
            # 获取原始图片尺寸
            original_width, original_height = img.size
            
            # 计算缩放比例
            width_ratio = fixed_width / original_width
            height_ratio = fixed_height / original_height
            scale = min(width_ratio, height_ratio)
            
            # 计算新的尺寸
            new_width = int(original_width * scale)
            new_height = int(original_height * scale)
            
            # 调整图片大小
            img = img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            # 将图片保存到BytesIO对象
            img_byte_arr = BytesIO()
            img.save(img_byte_arr, format=img.format if img.format else 'PNG')
            img_byte_arr.seek(0)
            
            # 创建Excel图片对象
            xl_img = XLImage(img_byte_arr)
            xl_img.anchor = cell_address
            worksheet.add_image(xl_img)
            
            return True
        except Exception as e:
            print(f"插入图片时发生错误: {str(e)}")
            return False

    def insert_product_image(self, worksheet, cell_address, msku, image_folder, fixed_width=None, fixed_height=None):
        """
        在Excel工作表中插入产品图片
        :param worksheet: openpyxl工作表对象
        :param cell_address: 单元格地址
        :param msku: 产品MSKU
        :param image_folder: 图片文件夹路径
        :param fixed_width: 固定宽度（可选）
        :param fixed_height: 固定高度（可选）
        """
        try:
            # 构建图片文件路径
            image_path_jpg = os.path.join(image_folder, f"{msku}.jpg")
            image_path_png = os.path.join(image_folder, f"{msku}.png")
            print(f"尝试加载图片: {image_path_jpg}")
            
            # 检查JPEG图片文件是否存在
            if os.path.exists(image_path_jpg):
                return self.insert_centered_image(worksheet, cell_address, image_path_jpg, fixed_width, fixed_height)
            elif os.path.exists(image_path_png):
                print(f"尝试加载PNG图片: {image_path_png}")
                return self.insert_centered_image(worksheet, cell_address, image_path_png, fixed_width, fixed_height)
            else:
                print(f"图片文件不存在: {image_path_jpg} 和 {image_path_png}")
                return False
        except Exception as e:
            print(f"处理产品图片时发生错误: {str(e)}")
            return False

    def insert_original_image(self, worksheet, cell_address, image_path):
        """
        在指定的单元格中插入原始图片，不进行压缩处理
        :param worksheet: 工作表对象
        :param cell_address: 单元格地址（例如'A1'）
        :param image_path: 图片文件路径
        :return: 是否成功插入图片
        """
        try:
            # 读取原始图片
            img = PILImage.open(image_path)
            
            # 将图片直接保存到BytesIO对象，不进行任何处理
            img_byte_arr = BytesIO()
            img.save(img_byte_arr, format=img.format if img.format else 'PNG')
            img_byte_arr.seek(0)
            
            # 创建Excel图片对象
            xl_img = XLImage(img_byte_arr)
            xl_img.anchor = cell_address
            worksheet.add_image(xl_img)
            
            print(f"成功插入原始图片，尺寸: {img.size}")
            return True
        except Exception as e:
            print(f"插入原始图片时发生错误: {str(e)}")
            return False

    # def insert_original_product_image(self, worksheet, cell_address, msku, image_folder):
        """
        在Excel工作表中插入原始产品图片，不进行压缩处理
        :param worksheet: openpyxl工作表对象
        :param cell_address: 单元格地址
        :param msku: 产品MSKU
        :param image_folder: 图片文件夹路径
        """
        try:
            # 构建图片文件路径
            image_path_jpg = os.path.join(image_folder, f"{msku}.jpg")
            image_path_png = os.path.join(image_folder, f"{msku}.png")
            print(f"尝试加载原始图片: {image_path_jpg}")
            
            # 检查JPEG图片文件是否存在
            if os.path.exists(image_path_jpg):
                return self.insert_original_image(worksheet, cell_address, image_path_jpg)
            elif os.path.exists(image_path_png):
                print(f"尝试加载PNG原始图片: {image_path_png}")
                return self.insert_original_image(worksheet, cell_address, image_path_png)
            else:
                print(f"图片文件不存在: {image_path_jpg} 和 {image_path_png}")
                return False
        except Exception as e:
            print(f"处理原始产品图片时发生错误: {str(e)}")
            return False
    def insert_original_product_image(self, worksheet, cell_address, msku, image_folder):
        """
        在Excel工作表中插入原始产品图片，不进行压缩处理
        :param worksheet: openpyxl工作表对象
        :param cell_address: 单元格地址
        :param msku: 产品MSKU
        :param image_folder: 图片文件夹路径
        """
        try:
            # 构建图片文件路径
            image_path_jpg = os.path.join(image_folder, f"{msku}.jpg")
            image_path_png = os.path.join(image_folder, f"{msku}.png")
            image_path_webp = os.path.join(image_folder, f"{msku}.webp")
            print(f"尝试加载原始图片: {image_path_jpg}")
            
            # 检查图片文件是否存在
            if os.path.exists(image_path_jpg):
                return self.insert_original_image(worksheet, cell_address, image_path_jpg)
            elif os.path.exists(image_path_png):
                print(f"尝试加载PNG原始图片: {image_path_png}")
                return self.insert_original_image(worksheet, cell_address, image_path_png)
            elif os.path.exists(image_path_webp):
                print(f"检测到WEBP格式图片，正在转换为PNG格式: {image_path_webp}")
                try:
                    from PIL import Image
                    # 生成转换后的PNG文件路径
                    converted_path = os.path.join(image_folder, f"{msku}_converted.png")
                    # 转换WEBP为PNG
                    with Image.open(image_path_webp) as img:
                        img.save(converted_path, 'PNG')
                    # 插入转换后的图片
                    result = self.insert_original_image(worksheet, cell_address, converted_path)
                    # 删除临时转换文件
                    if os.path.exists(converted_path):
                        os.remove(converted_path)
                    return result
                except ImportError:
                    print("警告：需要安装Pillow库来处理WEBP格式图片")
                    return False
                except Exception as e:
                    print(f"转换WEBP格式图片时发生错误: {str(e)}")
                    return False
            else:
                print(f"图片文件不存在: {image_path_jpg}、{image_path_png} 和 {image_path_webp}")
                return False
        except Exception as e:
            print(f"处理原始产品图片时发生错误: {str(e)}")
            return False
            
    def extract_data(self, ticket_str):
        import re
        # 定义正则表达式，修改为提取斜杠后的数字
        pattern = r'(?P<time>\d{4}\.\d{2}\.\d{2})-(?P<logistics>[^-]+(?:-[^-]+)*)-(?P<number>\d+)/(\d+)'
        match = re.search(pattern, ticket_str)
        if match:
            # 使用第四个捕获组（斜杠后的数字）作为件数
            return match.group('time'), match.group('logistics').replace('-', ''), match.group(4)
        else:
            return None, None, None

    def unmerge_cells_in_range(self, sheet, start_row, end_row, start_col, end_col):
        """
        解除指定范围内的所有合并单元格。

        :param sheet: 要操作的工作表对象
        :param start_row: 起始行
        :param end_row: 结束行
        :param start_col: 起始列
        :param end_col: 结束列
        """
        print("正在解除合并单元格...")
        merged_cells = list(sheet.merged_cells.ranges)
        cells_to_unmerge = []
        for merged_cell in merged_cells:
            min_row, min_col, max_row, max_col = merged_cell.bounds
            # 检查合并单元格是否在指定范围内
            if (min_row >= start_row and max_row <= end_row and
                min_col >= start_col and max_col <= end_col):
                try:
                    sheet.unmerge_cells(str(merged_cell))
                except Exception as e:
                    print(f"解除合并单元格时发生错误: {str(e)}")
        print("合并单元格解除完成")

    def merge_cells_in_range(self, sheet, start_row, end_row, start_col, end_col):
        """
        合并指定区域内的单元格。

        :param sheet: 要操作的工作表对象
        :param start_row: 起始行
        :param end_row: 结束行
        :param start_col: 起始列
        :param end_col: 结束列
        """
        try:
            # 获取合并区域的范围字符串
            merge_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            print(f"正在合并单元格区域: {merge_range}")
            
            # 合并单元格
            sheet.merge_cells(merge_range)
            
            # 设置合并后的单元格样式（可选）
            merged_cell = sheet.cell(row=start_row, column=start_col)
            merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"单元格合并完成: {merge_range}")
        except Exception as e:
            print(f"合并单元格时发生错误: {str(e)}")

    def _parse_range(self, range_str):
        """
        解析Excel单元格范围字符串

        :param range_str: Excel单元格范围字符串（例如'A1:B2'）
        :return: 解析后的范围元组（start_row, end_row, start_col, end_col）
        """
        start_cell, end_cell = range_str.split(':')
        start_row = int(start_cell[1:])
        start_col = ord(start_cell[0]) - 64
        end_row = int(end_cell[1:])
        end_col = ord(end_cell[0]) - 64
        return start_row, end_row, start_col, end_col

    def _create_declaration_table(self, sheet, row, product_info):
        """
        在指定行创建申报要素表格
        
        :param sheet: 工作表对象
        :param row: 起始行
        :param product_info: 产品信息
        :return: 表格占用的行数
        """
        # 创建表头
        cell = sheet.cell(row=row, column=2)  # 从第二列开始
        cell.value = "申报要素（必填）"
        cell.font = Font(name='SimSun', bold=True, size=11)
        cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)  # 合并第二列和第三列
        
        # 创建表格内容
        rows = [
            ("HS", product_info.get('hs_code', '')),
            ("品名", f"{product_info.get('en_name', '')}({product_info.get('cn_name', '')})"),
            ("材质", product_info.get('material_cn', '')),
            ("用途", product_info.get('usage_cn', '')),
            ("品牌", product_info.get('brand', '')),
            ("型号", product_info.get('model', '')),
        ]
        
        print(product_info.get('usage_cn', ''))
        for row_index, (label, value) in enumerate(rows, 1):
            # 标签列
            cell = sheet.cell(row=row + row_index, column=2)  # 第二列
            cell.value = label
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 值列
            cell = sheet.cell(row=row + row_index, column=3)  # 第三列
            cell.value = value
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
        
        return 7  # 返回表格占用的行数（1行表头 + 6行内容）
